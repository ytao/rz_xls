import openpyxl
# import datetime
# import time
import os
import conf as c


class project_data():
    def __init__(self):
        '''初始化'''
        self.filename=c.file_name
        if os.path.isfile(c.file_name)==False:
            self.init_success=False
        if os.path.isfile(c.file_name+'.lock')==True:
            self.init_success=False
        else:
            self.wb=openpyxl.load_workbook(self.filename)
            self.ws=self.wb[c.sheet_name]
            self.start_row=c.start_row
            self.max_row=c.max_row

            self.init_success=True


    def testWrite(self,income_information):
        '''测试功能'''
        for x in range(c.start_row,c.max_row):
            if not self.ws.cell(row=x,column=1).value:
                for i in income_information.keys():
                    self.ws.cell(row=x,column=c.dic_record_position[i]).value=income_information[i]
                self.wb.save(filename=self.filename)
                break
